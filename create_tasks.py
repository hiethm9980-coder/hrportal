from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "مهام المشروع"
ws.sheet_view.rightToLeft = True

# Colors
navy = "1A3A5C"
teal = "0D9488"
white = "FFFFFF"
light_gray = "F5F5F5"
light_blue = "E8F4FD"
light_green = "E8F5E9"
light_yellow = "FFF8E1"
light_orange = "FFF3E0"
light_purple = "F3E5F5"
light_red = "FFEBEE"
green_done = "4CAF50"
orange_progress = "FF9800"
blue_status = "2196F3"

# Column widths
widths = {"A": 8, "B": 45, "C": 12, "D": 22, "E": 12, "F": 14, "G": 35}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Styles
header_font = Font(name="Arial", bold=True, color=white, size=12)
header_fill = PatternFill("solid", fgColor=navy)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

main_font = Font(name="Arial", bold=True, color=white, size=11)
main_fill = PatternFill("solid", fgColor=teal)
main_align = Alignment(horizontal="right", vertical="center", wrap_text=True)

sub_font = Font(name="Arial", size=10)
sub_align = Alignment(horizontal="right", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Headers
headers = ["الرقم", "المهمة", "النوع", "المسؤول", "الأولوية", "الحالة", "الوصف"]
ws.append(headers)
for col in range(1, 8):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
ws.row_dimensions[1].height = 35

# Task data: (num, name, type, responsible, priority, status, description)
tasks = [
    # === 1. إعداد المشروع ===
    ("1", "إعداد بيئة التطوير والمشروع", "رئيسية", "مصطفى الأهدل", "عالية", "مكتمل", "تهيئة بيئة التطوير وإنشاء هيكل المشروع"),
    ("1.1", "إنشاء مشروع Flutter وتهيئة الهيكل", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "flutter create مع Clean Architecture"),
    ("1.2", "إعداد Firebase (Android/iOS/Web)", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "flutterfire configure لجميع المنصات"),
    ("1.3", "إعداد الحزم والتبعيات (pubspec.yaml)", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "Dio, Riverpod, GoRouter, GetIt, Firebase"),
    ("1.4", "إعداد نظام التوجيه (GoRouter)", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "تعريف المسارات وحماية الصفحات"),
    ("1.5", "إعداد حقن التبعيات (GetIt)", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "تسجيل ApiClient و Repositories"),

    # === 2. الشبكة والمصادقة ===
    ("2", "طبقة الشبكة والمصادقة", "رئيسية", "مصطفى الأهدل", "عالية", "مكتمل", "إعداد الاتصال بالسيرفر ونظام تسجيل الدخول"),
    ("2.1", "إعداد ApiClient مع Dio و Interceptors", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "AuthInterceptor + Debug Logging Interceptor"),
    ("2.2", "تطوير شاشة تسجيل الدخول", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "username/password مع validation"),
    ("2.3", "إدارة الجلسة (Token + Secure Storage)", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "حفظ واسترجاع التوكن و is_manager"),
    ("2.4", "شاشة البداية (Splash Screen)", "فرعية", "مصطفى الأهدل", "متوسطة", "مكتمل", "التحقق من الجلسة والتوجيه التلقائي"),
    ("2.5", "إرسال FCM Token مع تسجيل الدخول", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "FirebaseMessaging.instance.getToken()"),
    ("2.6", "تسجيل الخروج وتنظيف الجلسة", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "invalidate جميع Providers + مسح التخزين"),

    # === 3. واجهات المستخدم الأساسية ===
    ("3", "تطوير واجهات المستخدم الأساسية", "رئيسية", "مصطفى الأهدل", "عالية", "مكتمل", "تصميم وبرمجة جميع شاشات التطبيق"),
    ("3.1", "لوحة التحكم (Dashboard)", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "بطاقات إحصائية + إجراءات سريعة + تحية"),
    ("3.2", "شاشة الحضور والانصراف", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "كرت مخصص: ساعات العمل، التاريخ، وقت الدخول/الخروج"),
    ("3.3", "شاشة الإجازات + تقديم إجازة", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "عرض الإجازات + نموذج تقديم مع validation"),
    ("3.4", "شاشة الطلبات الإدارية + تقديم طلب", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "أنواع طلبات متعددة + حقل المبلغ للسلف"),
    ("3.5", "شاشة كشوف الرواتب", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "كرت الراتب مع التفاصيل (إجمالي، خصومات، صافي)"),
    ("3.6", "شاشة الملف الشخصي", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "هيدر متدرج + معلومات شخصية/وظيفية + طوارئ"),
    ("3.7", "شاشة تغيير كلمة المرور", "فرعية", "مصطفى الأهدل", "متوسطة", "مكتمل", "POST /api/v1/change-password"),
    ("3.8", "شاشة الإشعارات", "فرعية", "مصطفى الأهدل", "متوسطة", "مكتمل", "عرض الإشعارات الواردة"),
    ("3.9", "شاشة الإعدادات", "فرعية", "مصطفى الأهدل", "متوسطة", "مكتمل", "تغيير اللغة، الوضع الداكن، إعدادات عامة"),

    # === 4. نظام المدير ===
    ("4", "نظام اعتماد المدير", "رئيسية", "مصطفى الأهدل", "عالية", "مكتمل", "واجهة المدير لمراجعة واعتماد الطلبات"),
    ("4.1", "شاشة طلبات المدير (TabBar: إجازات | طلبات)", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "عرض طلبات الموظفين مع فلاتر الحالة"),
    ("4.2", "اعتماد/رفض الإجازات", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "Bottom Sheet مع تفاصيل + أزرار القبول/الرفض"),
    ("4.3", "اعتماد/رفض الطلبات الإدارية", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "نفس نمط الإجازات"),
    ("4.4", "حفظ صلاحية is_manager في Secure Storage", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "استمرار الصلاحية بعد إعادة التشغيل"),

    # === 5. التصميم والثيم ===
    ("5", "التصميم وتجربة المستخدم (UI/UX)", "رئيسية", "مصطفى الأهدل", "عالية", "مكتمل", "Material 3 + دعم RTL + وضع داكن"),
    ("5.1", "تطبيق Material 3 Theme (فاتح + داكن)", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "ألوان، أشكال، ظلال متناسقة"),
    ("5.2", "الانتقال من Google Fonts إلى خطوط محلية (Cairo)", "فرعية", "مصطفى الأهدل", "متوسطة", "مكتمل", "8 ملفات TTF محلية بدلاً من الإنترنت"),
    ("5.3", "دعم اللغة العربية (RTL) والترجمة", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "ملفات i18n: ar.json + en.json"),
    ("5.4", "تنسيق التاريخ والوقت مع دعم عربي (ص/م)", "فرعية", "مصطفى الأهدل", "متوسطة", "مكتمل", "AppFuns.formatApiDateTime() + formatDate()"),
    ("5.5", "تصميم كروت مخصصة لجميع الشاشات", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "حضور، إجازات، طلبات، رواتب"),
    ("5.6", "شارات الحالة (Status Badges) موحدة", "فرعية", "مصطفى الأهدل", "متوسطة", "مكتمل", "ألوان موحدة حسب الحالة في نهاية الصف"),

    # === 6. الإشعارات ===
    ("6", "نظام الإشعارات", "رئيسية", "مصطفى الأهدل", "عالية", "مكتمل", "Firebase Cloud Messaging + Awesome Notifications"),
    ("6.1", "إعداد Firebase Cloud Messaging", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "تهيئة FCM للاستقبال"),
    ("6.2", "إعداد Awesome Notifications للإشعارات المحلية", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "عرض الإشعارات عند وصولها"),

    # === 7. النشر - Android ===
    ("7", "نشر التطبيق على Google Play Store", "رئيسية", "مصطفى الأهدل", "عالية", "قيد المراجعة", "تجهيز ورفع التطبيق على متجر Google Play"),
    ("7.1", "إنشاء Upload Keystore وتوقيع التطبيق", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "keytool + key.properties + build.gradle.kts"),
    ("7.2", "بناء App Bundle (.aab)", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "flutter build appbundle --release"),
    ("7.3", "إعداد اسم التطبيق متعدد اللغات", "فرعية", "مصطفى الأهدل", "متوسطة", "مكتمل", "strings.xml: NH Employees / موظفين نيوهورايزن"),
    ("7.4", "إعداد بطاقة بيانات المتجر", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "وصف، لقطات شاشة، أيقونة، رسم مميز"),
    ("7.5", "إعداد سياسة الخصوصية", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "صفحة HTML عربي/إنجليزي"),
    ("7.6", "إكمال محتوى التطبيق (تقييم، إعلانات، أمان)", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "تقييم المحتوى + أمان البيانات + الإعلانات"),
    ("7.7", "إرسال التطبيق للمراجعة", "فرعية", "مصطفى الأهدل", "عالية", "قيد المراجعة", "بانتظار مراجعة Google (تم الإرسال 29 مارس 2026)"),

    # === 8. النشر - iOS ===
    ("8", "نشر التطبيق على Apple App Store", "رئيسية", "مصطفى الأهدل", "عالية", "قيد التنفيذ", "تجهيز ورفع التطبيق على متجر Apple"),
    ("8.1", "إعداد بيئة البناء على Mac (Flutter + Xcode)", "فرعية", "مصطفى الأهدل", "عالية", "مكتمل", "Flutter 3.41.6 + Xcode 26.1.1"),
    ("8.2", "نسخ المشروع وتشغيله على iOS Simulator", "فرعية", "مصطفى الأهدل", "عالية", "قيد التنفيذ", "حل مشاكل التوافق مع Xcode الجديد"),
    ("8.3", "إعداد حساب Apple Developer", "فرعية", "مصطفى الأهدل", "عالية", "لم يبدأ", "اشتراك $99/سنة + شهادات التوقيع"),
    ("8.4", "إعداد Firebase لـ iOS (GoogleService-Info.plist)", "فرعية", "مصطفى الأهدل", "عالية", "لم يبدأ", "تهيئة Firebase لمنصة iOS"),
    ("8.5", "بناء ملف IPA للإصدار", "فرعية", "مصطفى الأهدل", "عالية", "لم يبدأ", "flutter build ipa --release"),
    ("8.6", "رفع التطبيق على App Store Connect", "فرعية", "مصطفى الأهدل", "عالية", "لم يبدأ", "Xcode أو xcrun altool"),
    ("8.7", "إعداد بطاقة بيانات المتجر + المراجعة", "فرعية", "مصطفى الأهدل", "عالية", "لم يبدأ", "وصف، لقطات، تقييم، سياسة خصوصية"),
]

# Status colors
status_colors = {
    "مكتمل": "C8E6C9",
    "قيد التنفيذ": "FFF9C4",
    "قيد المراجعة": "BBDEFB",
    "لم يبدأ": "FFCDD2",
}

status_font_colors = {
    "مكتمل": "2E7D32",
    "قيد التنفيذ": "F57F17",
    "قيد المراجعة": "1565C0",
    "لم يبدأ": "C62828",
}

row_fills = {
    "رئيسية": PatternFill("solid", fgColor=teal),
    "فرعية_odd": PatternFill("solid", fgColor=light_gray),
    "فرعية_even": PatternFill("solid", fgColor=white),
}

for i, (num, name, typ, responsible, priority, status, desc) in enumerate(tasks):
    row = i + 2
    ws.cell(row=row, column=1, value=num)
    ws.cell(row=row, column=2, value=name)
    ws.cell(row=row, column=3, value=typ)
    ws.cell(row=row, column=4, value=responsible)
    ws.cell(row=row, column=5, value=priority)
    ws.cell(row=row, column=6, value=status)
    ws.cell(row=row, column=7, value=desc)

    is_main = typ == "رئيسية"

    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="right" if col != 1 else "center", vertical="center", wrap_text=True)

        if is_main:
            cell.font = Font(name="Arial", bold=True, color=white, size=11)
            cell.fill = main_fill
        else:
            fill_key = "فرعية_odd" if i % 2 == 0 else "فرعية_even"
            cell.font = Font(name="Arial", size=10)
            cell.fill = row_fills[fill_key]

    # Status cell styling
    status_cell = ws.cell(row=row, column=6)
    if not is_main and status in status_colors:
        status_cell.fill = PatternFill("solid", fgColor=status_colors[status])
        status_cell.font = Font(name="Arial", bold=True, size=10, color=status_font_colors[status])
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

    if is_main:
        ws.row_dimensions[row].height = 30
    else:
        ws.row_dimensions[row].height = 25

# Summary section
summary_row = len(tasks) + 3
ws.cell(row=summary_row, column=1, value="ملخص المشروع").font = Font(name="Arial", bold=True, size=13, color=navy)
ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=7)
ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal="center")

summaries = [
    ("اسم المشروع", "بوابة الموظفين - NH Employees"),
    ("المسؤول", "مصطفى الأهدل"),
    ("التقنيات", "Flutter + Dart + Firebase + Dio + Riverpod"),
    ("المنصات", "Android + iOS + Web"),
    ("إجمالي المهام الرئيسية", "8"),
    ("إجمالي المهام الفرعية", str(len([t for t in tasks if t[2] == "فرعية"]))),
    ("المكتملة", str(len([t for t in tasks if t[5] == "مكتمل"]))),
    ("قيد التنفيذ/المراجعة", str(len([t for t in tasks if t[5] in ("قيد التنفيذ", "قيد المراجعة")]))),
    ("لم تبدأ", str(len([t for t in tasks if t[5] == "لم يبدأ"]))),
]

for j, (label, value) in enumerate(summaries):
    r = summary_row + 1 + j
    ws.cell(row=r, column=2, value=label).font = Font(name="Arial", bold=True, size=10, color=navy)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=r, column=3, value=value).font = Font(name="Arial", size=10)
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)

output = r"C:\Users\mustapha\Documents\app_project\hr_portal\hr_portal\HR_Portal_Tasks.xlsx"
wb.save(output)
print(f"Saved: {output}")
